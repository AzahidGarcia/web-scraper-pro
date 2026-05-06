# Copyright (c) 2026 Adrian Azahid García / Strivark — MIT License
"""Tests for web_scraper.exporter — generate Excel/CSV, verify headers + data."""

from __future__ import annotations

import csv
from typing import TYPE_CHECKING

import openpyxl
import pytest

from web_scraper.exporter import export, export_csv, export_excel

if TYPE_CHECKING:
    from pathlib import Path

SAMPLE_RECORDS = [
    {"title": "Cozy Studio Downtown", "price": "$1,200", "url": "/listing/1"},
    {"title": "Spacious 2BR Suburb", "price": "$2,400", "url": "/listing/2"},
    {"title": "Modern Loft Midtown", "price": "$1,850", "url": "/listing/3"},
]


class TestExportExcel:
    def test_creates_file(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.xlsx"
        result = export_excel(SAMPLE_RECORDS, dest)
        assert result.exists()

    def test_has_header_row(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.xlsx"
        export_excel(SAMPLE_RECORDS, dest)
        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]  # type: ignore[union-attr]
        assert headers == ["title", "price", "url"]

    def test_data_rows(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.xlsx"
        export_excel(SAMPLE_RECORDS, dest)
        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        # Row 2 is first data row
        title_cell = ws.cell(row=2, column=1).value  # type: ignore[union-attr]
        assert title_cell == "Cozy Studio Downtown"

    def test_correct_row_count(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.xlsx"
        export_excel(SAMPLE_RECORDS, dest)
        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        # 1 header + 3 data rows
        assert ws.max_row == 4  # type: ignore[union-attr]

    def test_header_is_bold(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.xlsx"
        export_excel(SAMPLE_RECORDS, dest)
        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        cell = ws.cell(row=1, column=1)  # type: ignore[union-attr]
        assert cell.font.bold is True  # type: ignore[union-attr]

    def test_empty_records_creates_empty_file(self, tmp_path: Path) -> None:
        dest = tmp_path / "empty.xlsx"
        export_excel([], dest)
        assert dest.exists()

    def test_creates_parent_dirs(self, tmp_path: Path) -> None:
        dest = tmp_path / "subdir" / "nested" / "out.xlsx"
        export_excel(SAMPLE_RECORDS, dest)
        assert dest.exists()

    def test_currency_column_has_number_format(self, tmp_path: Path) -> None:
        dest = tmp_path / "currency.xlsx"
        records = [{"title": "Item A", "price": "1200.00"}]
        export_excel(records, dest)
        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        # The price cell (row 2, col 2) should have a currency number format
        price_cell = ws.cell(row=2, column=2)  # type: ignore[union-attr]
        assert "#,##0" in (price_cell.number_format or "")  # type: ignore[union-attr]


class TestExportCsv:
    def test_creates_file(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        result = export_csv(SAMPLE_RECORDS, dest)
        assert result.exists()

    def test_has_header_row(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        export_csv(SAMPLE_RECORDS, dest)
        with dest.open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            assert reader.fieldnames == ["title", "price", "url"]

    def test_correct_row_count(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        export_csv(SAMPLE_RECORDS, dest)
        with dest.open(encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 3

    def test_data_values(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        export_csv(SAMPLE_RECORDS, dest)
        with dest.open(encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["title"] == "Cozy Studio Downtown"
        assert rows[1]["price"] == "$2,400"

    def test_empty_records_creates_empty_file(self, tmp_path: Path) -> None:
        dest = tmp_path / "empty.csv"
        export_csv([], dest)
        assert dest.exists()
        assert dest.read_text(encoding="utf-8") == ""


class TestExportAutoFormat:
    def test_xlsx_extension_triggers_excel(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.xlsx"
        export(SAMPLE_RECORDS, dest)
        wb = openpyxl.load_workbook(dest)
        assert wb.active is not None  # type: ignore[union-attr]

    def test_csv_extension_triggers_csv(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.csv"
        export(SAMPLE_RECORDS, dest)
        with dest.open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 3

    def test_unsupported_format_raises(self, tmp_path: Path) -> None:
        dest = tmp_path / "out.json"
        with pytest.raises(ValueError, match="Unsupported output format"):
            export(SAMPLE_RECORDS, dest)
