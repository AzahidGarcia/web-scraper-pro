# Copyright (c) 2026 Adrian Azahid García / Strivark — MIT License
"""Export scraped data to Excel (.xlsx) or CSV files."""

from __future__ import annotations

import contextlib
import csv
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# ──────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ──────────────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_CURRENCY_RE_PATTERNS = ("price", "cost", "amount", "fee", "rate", "total")


def _looks_like_currency(col_name: str) -> bool:
    lower = col_name.lower()
    return any(p in lower for p in _CURRENCY_RE_PATTERNS)


def _auto_column_width(ws: Worksheet) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column or 1)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


# ──────────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────────

def export_excel(
    records: list[dict[str, Any]],
    path: str | Path,
    sheet_name: str = "Listings",
) -> Path:
    """Write *records* to an Excel workbook at *path*.

    The sheet has:
    - Bold, dark-blue header row with white text
    - Auto-fitted column widths
    - Number format for currency-looking columns
    - Frozen top row

    Parameters
    ----------
    records:
        List of dicts — each dict becomes a row; keys become headers.
    path:
        Destination file path (``*.xlsx``).
    sheet_name:
        Name of the worksheet (default ``"Listings"``).

    Returns
    -------
    Path
        The resolved path of the written file.
    """
    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = sheet_name

    if not records:
        wb.save(dest)
        return dest.resolve()

    headers = list(records[0].keys())

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value: Any = record.get(header, "")
            # Attempt numeric conversion
            if isinstance(value, str):
                stripped = value.replace(",", "").strip().lstrip("$€£¥")
                with contextlib.suppress(ValueError):
                    value = float(stripped)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, float) and _looks_like_currency(header):
                cell.number_format = '#,##0.00'

    # Freeze header row
    ws.freeze_panes = "A2"

    _auto_column_width(ws)

    wb.save(dest)
    return dest.resolve()


def export_csv(
    records: list[dict[str, Any]],
    path: str | Path,
) -> Path:
    """Write *records* to a CSV file at *path*.

    Parameters
    ----------
    records:
        List of dicts.
    path:
        Destination file path (``*.csv``).

    Returns
    -------
    Path
        The resolved path of the written file.
    """
    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)

    if not records:
        dest.write_text("", encoding="utf-8")
        return dest.resolve()

    headers = list(records[0].keys())
    with dest.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(records)

    return dest.resolve()


def export(
    records: list[dict[str, Any]],
    path: str | Path,
    fmt: str | None = None,
) -> Path:
    """Auto-detect format from file extension and export.

    Parameters
    ----------
    records:
        List of dicts.
    path:
        Destination path — extension determines format unless *fmt* is given.
    fmt:
        ``"xlsx"`` or ``"csv"`` — overrides extension detection.

    Returns
    -------
    Path
        Resolved path of the written file.
    """
    dest = Path(path)
    resolved_fmt = fmt or dest.suffix.lstrip(".").lower()
    if resolved_fmt == "xlsx":
        return export_excel(records, dest)
    if resolved_fmt == "csv":
        return export_csv(records, dest)
    raise ValueError(f"Unsupported output format '{resolved_fmt}'. Use 'xlsx' or 'csv'.")
